from flask import Flask, request, jsonify, g, send_from_directory
from flask_jwt_extended import JWTManager, jwt_required, create_access_token, get_jwt_identity
from flask_cors import CORS
from werkzeug.security import check_password_hash, generate_password_hash
import sqlite3
import os
from datetime import timedelta
import datetime
from werkzeug.utils import secure_filename
import pandas as pd
import openpyxl

app = Flask(__name__)

# 创建上传文件目录
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['JWT_SECRET_KEY'] = 'your-secret-key'  # 在生产环境中使用更安全的密钥
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(days=1)  # Token有效期1天
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
jwt = JWTManager(app)

# 数据库配置
DATABASE = 'database.sqlite'

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(error):
    if hasattr(g, 'db'):
        g.db.close()

# 创建测试用户和初始化数据库
def init_db():
    print("开始初始化数据库...")
    try:
        db = get_db()
        cursor = db.cursor()
        
        # 创建users表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                role TEXT NOT NULL
            )
        ''')
        print("users表创建成功")
        
        # 创建students表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS students (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                student_id TEXT UNIQUE NOT NULL,
                grade TEXT NOT NULL,
                class TEXT,
                photo_url TEXT,
                address TEXT,
                emergency_contact TEXT,
                emergency_phone TEXT,
                notes TEXT
            )
        ''')
        print("students表创建成功")
        
        # 创建behavior_types表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS behavior_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                category TEXT NOT NULL,
                description TEXT
            )
        ''')
        print("behavior_types表创建成功")
        
        # 创建behaviors表
        cursor.execute(''' SELECT count(name) FROM sqlite_master WHERE type='table' AND name='behaviors' ''')
        if cursor.fetchone()[0] == 0:
            print("创建behaviors表...")
            cursor.execute('''
                CREATE TABLE behaviors (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    student_id INTEGER NOT NULL,
                    behavior_type TEXT NOT NULL,
                    description TEXT,
                    date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    image_url TEXT,
                    FOREIGN KEY (student_id) REFERENCES students (id),
                    FOREIGN KEY (behavior_type) REFERENCES behavior_types (name)
                )
            ''')
            print("behaviors表创建成功")
            
            # 添加测试行为记录
            test_behaviors = [
                (1, '帮助同学', '主动帮助同学复习功课', None),
                (2, '迟到', '上午第一节课迟到5分钟', None),
                (3, '获奖', '在数学竞赛中获得一等奖', None)
            ]
            cursor.executemany('''
                INSERT INTO behaviors (student_id, behavior_type, description, image_url)
                VALUES (?, ?, ?, ?)
            ''', test_behaviors)
        
        # 检查是否需要插入测试数据
        cursor.execute('SELECT COUNT(*) FROM users')
        if cursor.fetchone()[0] == 0:
            print("插入测试用户数据...")
            # 创建测试用户
            test_password = 'admin123'
            hashed_password = generate_password_hash(test_password)
            cursor.execute('''
                INSERT INTO users (username, password, role)
                VALUES (?, ?, ?)
            ''', ('admin', hashed_password, 'admin'))
            
            print("插入测试学生数据...")
            # 添加测试学生数据
            test_students = [
                ('张三', 'S001', '高一', '1班', None, '北京市海淀区', '张父', '13800138000', '品学兼优'),
                ('李四', 'S002', '高二', '2班', None, '北京市朝阳区', '李母', '13900139000', '积极向上'),
                ('王五', 'S003', '高三', '3班', None, '北京市西城区', '王父', '13700137000', '认真负责')
            ]
            cursor.executemany('''
                INSERT INTO students (name, student_id, grade, class, photo_url, address, emergency_contact, emergency_phone, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', test_students)
            
            print("插入测试行为类型数据...")
            # 添加测试行为类型数据
            test_types = [
                ('迟到', '违纪', '上课迟到'),
                ('早退', '违纪', '未经许可提前离开'),
                ('打架', '违纪', '与他人发生肢体冲突'),
                ('帮助同学', '优秀', '主动帮助有困难的同学'),
                ('志愿服务', '优秀', '参与学校志愿服务活动'),
                ('获奖', '优秀', '在比赛或竞赛中获奖')
            ]
            cursor.executemany('''
                INSERT INTO behavior_types (name, category, description)
                VALUES (?, ?, ?)
            ''', test_types)
        
        db.commit()
        print("数据库初始化完成")
        
    except Exception as e:
        print("数据库初始化失败:", str(e))
        import traceback
        traceback.print_exc()
        if db:
            db.rollback()
        raise e

# 初始化数据库
with app.app_context():
    # 删除旧的数据库文件（如果存在）
    if os.path.exists(DATABASE):
        os.remove(DATABASE)
        print(f"已删除旧的数据库文件: {DATABASE}")
    
    init_db()

# 测试路由
@app.route('/api/test', methods=['GET'])
def test():
    return jsonify({'message': 'API is working!'})

# 登录路由
@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        print('收到登录请求数据:', data)  # 调试日志
        
        if not data:
            print('请求数据为空')  # 调试日志
            return jsonify({'message': '无效的请求数据'}), 400
            
        username = data.get('username')
        password = data.get('password')
        
        print(f'尝试登录: 用户名={username}, 密码长度={len(password) if password else 0}')  # 调试日志
        
        if not username or not password:
            print('用户名或密码为空')  # 调试日志
            return jsonify({'message': '用户名和密码不能为空'}), 400
        
        cur = get_db().cursor()
        user = cur.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        
        if user:
            print(f'找到用户: {dict(user)}')  # 调试日志
            stored_password = user['password']
            print(f'存储的密码哈希: {stored_password}')  # 调试日志
            
            is_valid = check_password_hash(stored_password, password)
            print(f'密码验证结果: {is_valid}')  # 调试日志
            
            if is_valid:
                access_token = create_access_token(identity=username)
                return jsonify({
                    'token': access_token,
                    'userInfo': {
                        'username': user['username'],
                        'role': user['role']
                    }
                })
        else:
            print(f'未找到用户: {username}')  # 调试日志
        
        return jsonify({'message': '用户名或密码错误'}), 401
            
    except Exception as e:
        print('登录失败:', str(e))
        import traceback
        traceback.print_exc()  # 打印详细的错误堆栈
        return jsonify({'message': '登录失败，请稍后重试'}), 500

@app.route('/api/verify-token', methods=['GET'])
@jwt_required()
def verify_token():
    try:
        current_user = get_jwt_identity()
        cur = get_db().cursor()
        user = cur.execute('SELECT username, role FROM users WHERE username = ?', (current_user,)).fetchone()
        if user:
            return jsonify({
                'valid': True,
                'userInfo': {
                    'username': user['username'],
                    'role': user['role']
                }
            })
        return jsonify({'valid': False}), 401
    except Exception as e:
        print('Token验证失败:', str(e))
        return jsonify({'valid': False}), 401

@app.route('/api/students', methods=['GET'])
@jwt_required()
def get_students():
    try:
        cur = get_db().cursor()
        students = cur.execute('''
            SELECT 
                s.*,
                (SELECT COUNT(*) FROM behaviors b 
                 JOIN behavior_types bt ON b.behavior_type = bt.name 
                 WHERE b.student_id = s.id AND bt.category = '违纪') as violation_count,
                (SELECT COUNT(*) FROM behaviors b 
                 JOIN behavior_types bt ON b.behavior_type = bt.name 
                 WHERE b.student_id = s.id AND bt.category = '优秀') as excellent_count
            FROM students s
        ''').fetchall()
        return jsonify([dict(row) for row in students])
    except Exception as e:
        print('获取学生列表失败:', str(e))
        return jsonify({'message': '获取学生列表失败'}), 500

@app.route('/api/students', methods=['POST'])
@jwt_required()
def add_student():
    try:
        data = request.json
        required_fields = ['name', 'student_id', 'grade']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'message': f'{field} 是必填项'}), 400
        
        cur = get_db().cursor()
        cur.execute('''
            INSERT INTO students (name, student_id, grade, class, photo_url, address, emergency_contact, emergency_phone, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['name'],
            data['student_id'],
            data['grade'],
            data.get('class'),
            data.get('photo_url'),
            data.get('address'),
            data.get('emergency_contact'),
            data.get('emergency_phone'),
            data.get('notes')
        ))
        get_db().commit()
        
        # 返回新创建的学生信息
        new_student = cur.execute('SELECT * FROM students WHERE id = ?', (cur.lastrowid,)).fetchone()
        return jsonify(new_student)
    except Exception as e:
        print('添加学生失败:', str(e))
        return jsonify({'message': '添加学生失败'}), 500

@app.route('/api/students/<int:id>', methods=['PUT'])
@jwt_required()
def update_student(id):
    try:
        data = request.json
        required_fields = ['name', 'student_id', 'grade']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'message': f'{field} 是必填项'}), 400
        
        cur = get_db().cursor()
        cur.execute('''
            UPDATE students 
            SET name = ?, student_id = ?, grade = ?, class = ?, photo_url = ?, 
                address = ?, emergency_contact = ?, emergency_phone = ?, notes = ?
            WHERE id = ?
        ''', (
            data['name'],
            data['student_id'],
            data['grade'],
            data.get('class'),
            data.get('photo_url'),
            data.get('address'),
            data.get('emergency_contact'),
            data.get('emergency_phone'),
            data.get('notes'),
            id
        ))
        get_db().commit()
        
        # 返回更新后的学生信息
        updated_student = cur.execute('SELECT * FROM students WHERE id = ?', (id,)).fetchone()
        return jsonify(updated_student)
    except Exception as e:
        print('更新学生信息失败:', str(e))
        return jsonify({'message': '更新学生信息失败'}), 500

@app.route('/api/students/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_student(id):
    try:
        cur = get_db().cursor()
        cur.execute('DELETE FROM students WHERE id = ?', (id,))
        get_db().commit()
        return jsonify({'message': '删除成功'})
    except Exception as e:
        print('删除学生失败:', str(e))
        return jsonify({'message': '删除学生失败'}), 500

@app.route('/api/behaviors', methods=['GET'])
@jwt_required()
def get_behaviors():
    try:
        cur = get_db().cursor()
        behaviors = cur.execute('''
            SELECT b.*, s.name as student_name, s.grade, s.class
            FROM behaviors b
            JOIN students s ON b.student_id = s.id
            ORDER BY b.date DESC
        ''').fetchall()
        
        # 转换为列表并处理图片URL
        result = []
        for behavior in behaviors:
            behavior_dict = dict(behavior)
            if behavior_dict.get('image_url'):
                # 确保返回完整的URL路径
                if not behavior_dict['image_url'].startswith('/api/'):
                    behavior_dict['image_url'] = f"/api/uploads/{behavior_dict['image_url'].split('/')[-1]}"
            result.append(behavior_dict)
            
        return jsonify(result)
    except Exception as e:
        print('获取行为记录失败:', str(e))
        return jsonify({'message': '获取行为记录失败'}), 500

@app.route('/api/behaviors', methods=['POST'])
@jwt_required()
def add_behavior():
    print("\n=== 开始处理添加行为记录请求 ===")
    try:
        data = request.get_json()
        print("请求数据:", data)
        
        if not data:
            print("错误: 无效的请求数据")
            return jsonify({'message': '无效的请求数据'}), 400
            
        # 验证必填字段
        required_fields = ['student_id', 'behavior_type', 'description', 'date']
        for field in required_fields:
            if not data.get(field):
                print(f"错误: 缺少必填字段 {field}")
                return jsonify({'message': f'{field} 是必填项'}), 400
        
        # 转换student_id为整数
        try:
            student_id = int(data['student_id'])
            print(f"student_id转换成功: {student_id}")
        except (ValueError, TypeError) as e:
            print(f"错误: student_id转换失败 - {str(e)}")
            return jsonify({'message': '无效的学生ID'}), 400
            
        # 获取数据库连接
        db = get_db()
        cur = db.cursor()
        
        try:
            # 验证学生是否存在
            student = cur.execute('SELECT id FROM students WHERE id = ?', (student_id,)).fetchone()
            if not student:
                print(f"错误: 学生ID {student_id} 不存在")
                return jsonify({'message': '学生不存在'}), 400
            print(f"学生验证成功: {student_id}")
            
            # 验证行为类型是否存在
            behavior_type = cur.execute('SELECT name FROM behavior_types WHERE name = ?', 
                                      (data['behavior_type'],)).fetchone()
            if not behavior_type:
                print(f"错误: 行为类型 {data['behavior_type']} 不存在")
                return jsonify({'message': '行为类型不存在'}), 400
            print(f"行为类型验证成功: {data['behavior_type']}")
            
            # 插入行为记录
            print("开始插入行为记录...")
            cur.execute('''
                INSERT INTO behaviors (student_id, behavior_type, description, date, image_url)
                VALUES (?, ?, ?, ?, ?)
            ''', (
                student_id,
                data['behavior_type'],
                data['description'],
                data['date'],
                data.get('image_url')
            ))
            db.commit()
            print("行为记录插入成功")
            
            # 获取新插入的记录
            print("开始获取新插入的记录...")
            new_behavior = cur.execute('''
                SELECT b.*, s.name as student_name, s.grade, s.class
                FROM behaviors b
                JOIN students s ON b.student_id = s.id
                WHERE b.id = last_insert_rowid()
            ''').fetchone()
            
            if not new_behavior:
                print("错误: 无法获取新创建的记录")
                return jsonify({'message': '记录创建失败'}), 500
                
            # 转换为字典并返回
            result = dict(new_behavior)
            print("成功获取新记录:", result)
            return jsonify(result)
            
        except sqlite3.Error as e:
            db.rollback()
            print(f"数据库错误: {str(e)}")
            return jsonify({'message': f'数据库操作失败: {str(e)}'}), 500
            
    except Exception as e:
        import traceback
        print(f"未知错误: {str(e)}")
        traceback.print_exc()
        return jsonify({'message': f'服务器内部错误: {str(e)}'}), 500
    finally:
        print("=== 处理添加行为记录请求结束 ===\n")

@app.route('/api/behaviors/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_behavior(id):
    try:
        cur = get_db().cursor()
        cur.execute('DELETE FROM behaviors WHERE id = ?', (id,))
        get_db().commit()
        return jsonify({'message': '删除成功'})
    except Exception as e:
        print('删除行为记录失败:', str(e))
        return jsonify({'message': '删除行为记录失败'}), 500

@app.route('/api/behaviors/<int:id>', methods=['PUT'])
@jwt_required()
def update_behavior(id):
    print(f"\n=== 开始更新行为记录 ID:{id} ===")
    try:
        data = request.get_json()
        if not data:
            print("错误: 无效的请求数据")
            return jsonify({'message': '无效的请求数据'}), 400
            
        # 验证必填字段
        required_fields = ['student_id', 'behavior_type', 'description', 'date']
        for field in required_fields:
            if not data.get(field):
                print(f"错误: 缺少必填字段 {field}")
                return jsonify({'message': f'{field} 是必填项'}), 400
        
        # 转换student_id为整数
        try:
            student_id = int(data['student_id'])
            print(f"student_id转换成功: {student_id}")
        except (ValueError, TypeError) as e:
            print(f"错误: student_id转换失败 - {str(e)}")
            return jsonify({'message': '无效的学生ID'}), 400
            
        # 获取数据库连接
        db = get_db()
        cur = db.cursor()
        
        try:
            # 验证行为记录是否存在
            behavior = cur.execute('SELECT id FROM behaviors WHERE id = ?', (id,)).fetchone()
            if not behavior:
                print(f"错误: 行为记录ID {id} 不存在")
                return jsonify({'message': '行为记录不存在'}), 404
                
            # 验证学生是否存在
            student = cur.execute('SELECT id FROM students WHERE id = ?', (student_id,)).fetchone()
            if not student:
                print(f"错误: 学生ID {student_id} 不存在")
                return jsonify({'message': '学生不存在'}), 400
                
            # 验证行为类型是否存在
            behavior_type = cur.execute('SELECT name FROM behavior_types WHERE name = ?', 
                                      (data['behavior_type'],)).fetchone()
            if not behavior_type:
                print(f"错误: 行为类型 {data['behavior_type']} 不存在")
                return jsonify({'message': '行为类型不存在'}), 400
            
            # 更新行为记录
            print("开始更新行为记录...")
            cur.execute('''
                UPDATE behaviors 
                SET student_id = ?, behavior_type = ?, description = ?, date = ?, image_url = ?
                WHERE id = ?
            ''', (
                student_id,
                data['behavior_type'],
                data['description'],
                data['date'],
                data.get('image_url'),
                id
            ))
            db.commit()
            print("行为记录更新成功")
            
            # 获取更新后的记录
            print("获取更新后的记录...")
            updated_behavior = cur.execute('''
                SELECT b.*, s.name as student_name, s.grade, s.class
                FROM behaviors b
                JOIN students s ON b.student_id = s.id
                WHERE b.id = ?
            ''', (id,)).fetchone()
            
            if not updated_behavior:
                print("错误: 无法获取更新后的记录")
                return jsonify({'message': '记录更新失败'}), 500
                
            # 转换为字典并返回
            result = dict(updated_behavior)
            print("成功获取更新后的记录:", result)
            return jsonify(result)
            
        except sqlite3.Error as e:
            db.rollback()
            print(f"数据库错误: {str(e)}")
            return jsonify({'message': f'数据库操作失败: {str(e)}'}), 500
            
    except Exception as e:
        import traceback
        print(f"未知错误: {str(e)}")
        traceback.print_exc()
        return jsonify({'message': f'服务器内部错误: {str(e)}'}), 500
    finally:
        print("=== 更新行为记录结束 ===\n")

@app.route('/api/behavior-types', methods=['GET'])
@jwt_required()
def get_behavior_types():
    try:
        cur = get_db().cursor()
        behavior_types = cur.execute('SELECT * FROM behavior_types').fetchall()
        return jsonify([dict(row) for row in behavior_types])
    except Exception as e:
        print('获取行为类型失败:', str(e))
        return jsonify({'message': '获取行为类型失败'}), 500

@app.route('/api/behavior-types', methods=['POST'])
@jwt_required()
def add_behavior_type():
    try:
        data = request.json
        required_fields = ['name', 'category', 'description']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'message': f'{field} 是必填项'}), 400
        
        cur = get_db().cursor()
        cur.execute('''
            INSERT INTO behavior_types (name, category, description)
            VALUES (?, ?, ?)
        ''', (
            data['name'],
            data['category'],
            data['description']
        ))
        get_db().commit()
        
        # 返回新创建的行为类型
        new_type = cur.execute('SELECT * FROM behavior_types WHERE id = ?', (cur.lastrowid,)).fetchone()
        return jsonify(new_type)
    except Exception as e:
        print('添加行为类型失败:', str(e))
        return jsonify({'message': '添加行为类型失败'}), 500

@app.route('/api/behavior-types/<int:id>', methods=['PUT'])
@jwt_required()
def update_behavior_type(id):
    try:
        data = request.json
        required_fields = ['name', 'category', 'description']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'message': f'{field} 是必填项'}), 400
        
        cur = get_db().cursor()
        cur.execute('''
            UPDATE behavior_types 
            SET name = ?, category = ?, description = ?
            WHERE id = ?
        ''', (
            data['name'],
            data['category'],
            data['description'],
            id
        ))
        get_db().commit()
        
        # 返回更新后的行为类型
        updated_type = cur.execute('SELECT * FROM behavior_types WHERE id = ?', (id,)).fetchone()
        return jsonify(updated_type)
    except Exception as e:
        print('更新行为类型失败:', str(e))
        return jsonify({'message': '更新行为类型失败'}), 500

@app.route('/api/behavior-types/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_behavior_type(id):
    try:
        cur = get_db().cursor()
        cur.execute('DELETE FROM behavior_types WHERE id = ?', (id,))
        get_db().commit()
        return jsonify({'message': '删除成功'})
    except Exception as e:
        print('删除行为类型失败:', str(e))
        return jsonify({'message': '删除行为类型失败'}), 500

@app.route('/api/statistics', methods=['GET'])
@jwt_required()
def get_statistics():
    try:
        # 获取查询参数
        grade = request.args.get('grade')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 构建基础查询
        behaviors_query = '''
            SELECT b.*, s.grade, s.name as student_name, bt.name as behavior_type, bt.category
            FROM behaviors b
            JOIN students s ON b.student_id = s.id
            JOIN behavior_types bt ON b.behavior_type = bt.name
        '''
        
        query_params = []
        where_conditions = []
        
        # 添加筛选条件
        if grade:
            where_conditions.append("s.grade = ?")
            query_params.append(grade)
        
        if start_date:
            where_conditions.append("b.date >= ?")
            query_params.append(start_date)
        
        if end_date:
            where_conditions.append("b.date <= ?")
            query_params.append(end_date)
        
        if where_conditions:
            behaviors_query += " WHERE " + " AND ".join(where_conditions)
        
        # 执行查询
        cur = get_db().cursor()
        behaviors = [dict(row) for row in cur.execute(behaviors_query, query_params).fetchall()]
        
        # 统计数据
        violation_behaviors = [b for b in behaviors if b['category'] == '违纪']
        excellent_behaviors = [b for b in behaviors if b['category'] == '优秀']
        
        # 获取违纪和优秀学生的唯一ID
        violation_students = len(set(b['student_id'] for b in violation_behaviors))
        excellent_students = len(set(b['student_id'] for b in excellent_behaviors))
        
        # 行为类型分布
        behavior_type_count = {}
        for behavior in behaviors:
            type_name = behavior['behavior_type']
            behavior_type_count[type_name] = behavior_type_count.get(type_name, 0) + 1
        
        behavior_type_distribution = [
            {'name': type_name, 'value': count}
            for type_name, count in behavior_type_count.items()
        ]
        
        # 年级行为趋势
        grades = ['高一', '高二', '高三']
        grade_violations = []
        grade_excellent = []
        
        for g in grades:
            grade_violations.append(len([b for b in violation_behaviors if b['grade'] == g]))
            grade_excellent.append(len([b for b in excellent_behaviors if b['grade'] == g]))
        
        # 时间趋势
        violation_trend = []
        excellent_trend = []
        
        if behaviors:
            # 按日期分组统计
            date_stats = {}
            for behavior in behaviors:
                date = behavior['date'].split('T')[0]  # 只取日期部分
                if date not in date_stats:
                    date_stats[date] = {'violations': 0, 'excellent': 0}
                
                if behavior['category'] == '违纪':
                    date_stats[date]['violations'] += 1
                else:
                    date_stats[date]['excellent'] += 1
            
            # 转换为时间序列数据
            for date in sorted(date_stats.keys()):
                violation_trend.append([date, date_stats[date]['violations']])
                excellent_trend.append([date, date_stats[date]['excellent']])
        
        return jsonify({
            'total_violations': len(violation_behaviors),
            'total_excellent': len(excellent_behaviors),
            'violation_students': violation_students,
            'excellent_students': excellent_students,
            'behavior_type_distribution': behavior_type_distribution,
            'grade_violations': grade_violations,
            'grade_excellent': grade_excellent,
            'time_trend': {
                'violations': violation_trend,
                'excellent': excellent_trend
            }
        })
        
    except Exception as e:
        print('统计数据获取失败:', str(e))
        return jsonify({'message': '获取统计数据失败'}), 500

# 文件上传路由
@app.route('/api/upload', methods=['POST'])
@jwt_required()
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'message': '没有文件被上传'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'message': '没有选择文件'}), 400
            
        if file:
            # 确保文件名安全
            filename = secure_filename(file.filename)
            # 生成唯一的文件名
            unique_filename = f"{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
            # 确保上传目录存在
            if not os.path.exists(app.config['UPLOAD_FOLDER']):
                os.makedirs(app.config['UPLOAD_FOLDER'])
            # 保存文件
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            file.save(file_path)
            # 返回完整的URL路径
            file_url = f'/api/uploads/{unique_filename}'
            return jsonify({'url': file_url})
    except Exception as e:
        print('文件上传失败:', str(e))
        return jsonify({'message': '文件上传失败'}), 500

# 提供上传文件的访问路由
@app.route('/api/uploads/<filename>')
def uploaded_file(filename):
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except Exception as e:
        print('文件访问失败:', str(e))
        return jsonify({'message': '文件不存在'}), 404

# 学生导入路由
@app.route('/api/students/import', methods=['POST'])
@jwt_required()
def import_students():
    try:
        if 'file' not in request.files:
            return jsonify({'message': '没有文件被上传'}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'message': '没有选择文件'}), 400
            
        if file and file.filename.endswith(('.xlsx', '.xls')):
            # 保存文件
            filename = secure_filename(file.filename)
            temp_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(temp_path)
            
            try:
                # 读取Excel文件
                df = pd.read_excel(temp_path)
                
                # 验证必填字段
                required_fields = ['学号', '姓名', '年级']
                for field in required_fields:
                    if field not in df.columns:
                        return jsonify({'message': f'缺少必填字段: {field}'}), 400
                
                # 开始导入数据
                success_count = 0
                error_count = 0
                error_messages = []
                
                cur = get_db().cursor()
                
                for index, row in df.iterrows():
                    try:
                        # 检查学号是否已存在
                        existing = cur.execute('SELECT id FROM students WHERE student_id = ?', 
                                            (str(row['学号']),)).fetchone()
                        if existing:
                            error_count += 1
                            error_messages.append(f'第{index+2}行: 学号 {row["学号"]} 已存在')
                            continue
                        
                        # 插入数据
                        cur.execute('''
                            INSERT INTO students (
                                name, student_id, grade, class, 
                                address, emergency_contact, emergency_phone
                            ) VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            str(row['姓名']),
                            str(row['学号']),
                            str(row['年级']),
                            str(row.get('班级', '')),
                            str(row.get('家庭住址', '')),
                            str(row.get('紧急联系人', '')),
                            str(row.get('联系人电话', ''))
                        ))
                        success_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        error_messages.append(f'第{index+2}行: {str(e)}')
                
                get_db().commit()
                
                # 删除临时文件
                os.remove(temp_path)
                
                return jsonify({
                    'message': '导入完成',
                    'success_count': success_count,
                    'error_count': error_count,
                    'error_messages': error_messages
                })
                
            except Exception as e:
                return jsonify({'message': f'解析Excel文件失败: {str(e)}'}), 400
            finally:
                # 确保临时文件被删除
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                    
        return jsonify({'message': '不支持的文件格式'}), 400
        
    except Exception as e:
        print('导入学生失败:', str(e))
        return jsonify({'message': '导入学生失败'}), 500

# 添加模板下载路由
@app.route('/api/students/template', methods=['GET'])
@jwt_required()
def download_template():
    print("\n=== 开始生成学生导入模板 ===")
    try:
        print("创建新的Excel工作簿...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "学生信息"

        print("添加表头...")
        headers = ['学号', '姓名', '年级', '班级', '家庭住址', '紧急联系人', '联系人电话', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        print("添加示例数据...")
        example_data = [
            ['S001', '张三', '高一', '1班', '北京市海淀区', '张父', '13800138000', '品学兼优'],
            ['S002', '李四', '高二', '2班', '北京市朝阳区', '李母', '13900139000', '积极向上']
        ]
        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        print("调整列宽...")
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        print("保存模板文件...")
        # 使用绝对路径
        temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp')
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)
            
        temp_file = os.path.join(temp_dir, 'student_import_template.xlsx')
        wb.save(temp_file)

        print("发送文件...")
        response = send_from_directory(temp_dir, 'student_import_template.xlsx', as_attachment=True)
        
        # 添加清理代码
        @response.call_on_close
        def cleanup():
            try:
                os.remove(temp_file)
                print(f"临时文件已删除: {temp_file}")
            except Exception as e:
                print(f"清理临时文件失败: {str(e)}")
                
        print("=== 模板生成完成 ===\n")
        return response
        
    except Exception as e:
        import traceback
        print("生成模板文件失败:")
        print(str(e))
        traceback.print_exc()
        return jsonify({
            'message': '生成模板文件失败',
            'error': str(e)
        }), 500

@app.route('/api/statistics/behavior-trends')
@jwt_required()
def get_behavior_trends():
    days = request.args.get('days', '7')
    days = int(days)
    
    db = get_db()
    cur = db.cursor()
    
    # 获取指定天数内的行为记录
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    
    # 生成日期列表
    dates = []
    violations = []
    excellents = []
    current_date = start_date
    
    while current_date <= end_date:
        date_str = current_date.strftime('%Y-%m-%d')
        dates.append(date_str)
        
        # 获取违纪行为数量
        cur.execute('''
            SELECT COUNT(*) FROM behaviors b
            JOIN behavior_types bt ON b.type_id = bt.id
            WHERE DATE(b.created_at) = ? AND bt.is_violation = 1
        ''', (date_str,))
        violations.append(cur.fetchone()[0])
        
        # 获取优秀表现数量
        cur.execute('''
            SELECT COUNT(*) FROM behaviors b
            JOIN behavior_types bt ON b.type_id = bt.id
            WHERE DATE(b.created_at) = ? AND bt.is_violation = 0
        ''', (date_str,))
        excellents.append(cur.fetchone()[0])
        
        current_date += timedelta(days=1)
    
    return jsonify({
        'dates': dates,
        'violations': violations,
        'excellents': excellents
    })

@app.route('/api/statistics/behavior-types')
@jwt_required()
def get_behavior_type_distribution():
    db = get_db()
    cur = db.cursor()
    
    cur.execute('''
        SELECT bt.name, bt.is_violation, COUNT(b.id) as count
        FROM behavior_types bt
        LEFT JOIN behaviors b ON bt.id = b.type_id
        GROUP BY bt.id
        ORDER BY count DESC
    ''')
    
    results = []
    for row in cur.fetchall():
        results.append({
            'type_name': row[0],
            'is_violation': bool(row[1]),
            'count': row[2]
        })
    
    return jsonify(results)

@app.route('/api/statistics/grade-comparison')
@jwt_required()
def get_grade_comparison():
    db = get_db()
    cur = db.cursor()
    
    grades = ['高一', '高二', '高三']
    violations = []
    excellents = []
    
    for grade in grades:
        # 获取违纪行为数量
        cur.execute('''
            SELECT COUNT(*) FROM behaviors b
            JOIN behavior_types bt ON b.type_id = bt.id
            JOIN students s ON b.student_id = s.id
            WHERE s.grade = ? AND bt.is_violation = 1
        ''', (grade,))
        violations.append(cur.fetchone()[0])
        
        # 获取优秀表现数量
        cur.execute('''
            SELECT COUNT(*) FROM behaviors b
            JOIN behavior_types bt ON b.type_id = bt.id
            JOIN students s ON b.student_id = s.id
            WHERE s.grade = ? AND bt.is_violation = 0
        ''', (grade,))
        excellents.append(cur.fetchone()[0])
    
    return jsonify({
        'violations': violations,
        'excellents': excellents
    })

if __name__ == '__main__':
    app.run(debug=True, port=5002, host='0.0.0.0') 